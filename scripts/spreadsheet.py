from openpyxl import Workbook
from openpyxl import load_workbook

def prepare_results_workbook(ques_filename):
    # define workbook
    res_book = Workbook()

    # select first sheet(active always returns first sheet)
    sheet1 = res_book.active
    
    # add the heading cols
    sheet1.cell(1, 1).value = "Email Address"
    sheet1.cell(1, 2).value = "No. Of Correct"

    # open questions sheet
    ques_sheet = load_workbook(ques_filename).active

    # find the no. of records
    rec = len(ques_sheet['B'])

    # copy email, and put 0 for no. of correct
    for row_iter in range(2, rec + 1):
        sheet1.cell(row_iter, 1).value = ques_sheet.cell(row_iter, 2).value
        sheet1.cell(row_iter, 2).value = 0

    # save workbook
    res_book.save("Results.xlsx")

def correct_answers(ques_filename, ans_filename):
    # open questions sheet
    ques_sheet = load_workbook(ques_filename).active

    # find the no. of records
    rec = len(ques_sheet['B'])

    # find no. of columns
    col = ques_sheet.max_column

    # open answers sheet
    ans_sheet = load_workbook(ans_filename).active

    # open results book and select first(active) sheet
    res_book = load_workbook("Results.xlsx")
    res_sheet = res_book.active

    # start correcting
    for i in range(2, rec+1):
        for j in range(3, col):
            counter = 0
            for k in range(1, ans_sheet.max_row):
                if ques_sheet.cell(i, j).value == ans_sheet.cell(k, j - 2).value:
                    counter += 1
            if counter > 0:
                res_sheet.cell(i, 2).value += 1
    
    # save results book
    res_book.save("Results.xlsx")

quesf = input()
ansf = input()

prepare_results_workbook(quesf)

correct_answers(quesf, ansf)
